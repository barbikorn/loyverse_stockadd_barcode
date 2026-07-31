"""
excel_export.py — สร้างไฟล์ Excel (.xlsx) จาก SalesReport

โครงสร้างไฟล์:
  ชีตแรก  "สรุปรวมทุกราย"  → ภาพรวมทุกผู้ฝากขาย (สำหรับใช้ภายใน)
  ชีตถัดไป 1 ชีตต่อผู้ฝากขาย 1 ราย → จัดหน้าแบบ "ใบสรุปยอดขายฝากขาย"
                                     (รายการสินค้า + จำนวน + ราคา/หน่วย + ยอดรวม + บล็อกสรุปยอด)
  ออกแบบให้ปริ้นต์หรือส่งต่อให้ผู้ฝากขายแต่ละรายได้เลย โดยไม่เห็นข้อมูลของรายอื่น
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from loyverse import config
from loyverse.reports.sales_by_category import SalesReport

BRAND = "5E2F10"

HEADER_FILL = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=16, color=BRAND)
SUBTITLE_FONT = Font(size=10, color="8A6D4F")
BOLD = Font(bold=True)
LABEL_FONT = Font(bold=True, color="8A6D4F", size=10)
PAYOUT_FONT = Font(bold=True, size=13, color=BRAND)

THIN = Side(style="thin", color="D9CBB3")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BORDER = Border(top=Side(style="thin"))
DOUBLE_TOP = Border(top=Side(style="double", color=BRAND))

MONEY_FMT = "#,##0.00"
QTY_FMT = "#,##0.###"
PCT_FMT = '0.00"%"'

RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

SUMMARY_HEADERS = [
    "ผู้ฝากขาย (Category)", "จำนวนที่ขาย", "ยอดขายรวม (Gross)", "ส่วนลด",
    "ยอดสุทธิที่ต้องจ่าย", "สัดส่วน %", "จำนวนบิล",
]

STATEMENT_HEADERS = [
    "ลำดับ", "SKU", "ชื่อสินค้า", "ตัวเลือก", "จำนวน",
    "ราคา/หน่วย", "ยอดขายรวม", "ส่วนลด", "ยอดสุทธิ",
]

# ความกว้างคอลัมน์ของชีต statement (A..I)
_STATEMENT_WIDTHS = [7, 14, 34, 14, 10, 13, 14, 12, 15]

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """ชื่อชีตที่ Excel ยอมรับ: ไม่เกิน 31 ตัว, ไม่มีอักขระต้องห้าม, ไม่ซ้ำกัน"""
    clean = _INVALID_SHEET_CHARS.sub("-", (name or "").strip()) or "ไม่ระบุ"
    clean = clean[:31]
    candidate = clean
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def suggested_filename(report: SalesReport) -> str:
    """
    ตั้งชื่อไฟล์ให้สื่อความ — ถ้ารายงานมีผู้ฝากขายรายเดียว ใส่ชื่อรายนั้นลงในชื่อไฟล์
    เพื่อให้ส่งไฟล์ให้เจ้าของได้เลยโดยไม่ต้องเปลี่ยนชื่อ
    """
    period = f"{report.date_from:%Y%m%d}-{report.date_to:%Y%m%d}"
    if len(report.rows) == 1:
        owner = re.sub(r'[\\/*?:"<>|]', "_", report.rows[0].category_name).strip()
        return f"สรุปยอดฝากขาย_{owner}_{period}.xlsx"
    return f"สรุปยอดฝากขาย_ทุกราย_{period}.xlsx"


def _residual(gross: float, discounts: float, net: float) -> float:
    """
    ส่วนต่างที่ net อธิบายไม่ได้ด้วย gross - discount (เช่นภาษี/ปรับปรุงอื่น)
    ปกติร้านที่ไม่มี VAT ค่านี้จะเป็น 0 — ถ้าไม่ใช่ 0 เราจะโชว์เป็นบรรทัดแยกไม่ให้ยอดดูขัดกัน
    """
    return net - (gross - discounts)


# ─── ชีตสรุปรวมทุกราย ──────────────────────────────────────────

def _write_overview_sheet(wb: Workbook, report: SalesReport) -> None:
    ws = wb.active
    ws.title = "สรุปรวมทุกราย"

    ws["A1"] = "สรุปยอดขายแยกตามผู้ฝากขาย"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUMMARY_HEADERS))

    cat_desc = ("ทุกราย" if not report.category_filter
                else f"{len(report.category_filter)} รายที่เลือก")
    info_lines = [
        f"ช่วงวันที่: {report.date_from} ถึง {report.date_to}  ({report.tz_name})",
        f"ขอบเขต: {cat_desc}",
        f"ออกรายงานเมื่อ: {report.generated_at:%Y-%m-%d %H:%M:%S} UTC",
        f"บิลที่นับ: {report.receipts_scanned} ใบ "
        f"(เป็นการคืนสินค้า {report.refunds_scanned} ใบ · บิลที่ยกเลิกและไม่นับ {report.cancelled_skipped} ใบ)",
    ]
    r = 2
    for line in info_lines:
        ws.cell(row=r, column=1, value=line).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(SUMMARY_HEADERS))
        r += 1

    header_row = r + 1
    for col, title in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BOX

    first = header_row + 1
    rr = first
    for row in report.rows:
        ws.cell(row=rr, column=1, value=row.category_name).border = BOX
        ws.cell(row=rr, column=2, value=row.qty_sold).number_format = QTY_FMT
        ws.cell(row=rr, column=3, value=row.gross_sales).number_format = MONEY_FMT
        ws.cell(row=rr, column=4, value=row.discounts).number_format = MONEY_FMT
        ws.cell(row=rr, column=5, value=row.net_sales).number_format = MONEY_FMT
        ws.cell(row=rr, column=6, value=row.share_pct).number_format = PCT_FMT
        ws.cell(row=rr, column=7, value=row.receipts_count)
        for c in range(2, len(SUMMARY_HEADERS) + 1):
            ws.cell(row=rr, column=c).border = BOX
        rr += 1
    last = rr - 1
    has_rows = last >= first

    total_row = rr
    ws.cell(row=total_row, column=1, value="รวมทั้งหมด").font = BOLD
    for col, fmt in ((2, QTY_FMT), (3, MONEY_FMT), (4, MONEY_FMT), (5, MONEY_FMT), (7, None)):
        letter = get_column_letter(col)
        value = f"=SUM({letter}{first}:{letter}{last})" if has_rows else 0
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = BOLD
        if fmt:
            cell.number_format = fmt
    pct = ws.cell(row=total_row, column=6, value=report.totals.share_pct)
    pct.font = BOLD
    pct.number_format = PCT_FMT
    for col in range(1, len(SUMMARY_HEADERS) + 1):
        ws.cell(row=total_row, column=col).border = DOUBLE_TOP

    ws.freeze_panes = ws.cell(row=first, column=1).coordinate
    if has_rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(SUMMARY_HEADERS))}{last}"

    ws.column_dimensions["A"].width = 34
    for col in range(2, len(SUMMARY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


# ─── ชีต statement ต่อผู้ฝากขาย 1 ราย ──────────────────────────

def _write_statement_sheet(wb: Workbook, report: SalesReport, cat_row, sheet_name: str,
                           doc_index: int) -> None:
    ws = wb.create_sheet(sheet_name)
    ncols = len(STATEMENT_HEADERS)
    last_col_letter = get_column_letter(ncols)

    for i, width in enumerate(_STATEMENT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── หัวเอกสาร ──
    ws["A1"] = "ใบสรุปยอดขายสินค้าฝากขาย"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Consignment Sales Statement"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    doc_no = f"CS-{report.date_from:%Y%m%d}-{report.date_to:%Y%m%d}-{doc_index:02d}"
    info_rows = [
        ("ผู้ฝากขาย", cat_row.category_name),
        ("ช่วงวันที่ขาย", f"{report.date_from} ถึง {report.date_to}  ({report.tz_name})"),
        ("วันที่ออกเอกสาร", f"{report.generated_at:%Y-%m-%d %H:%M:%S} UTC"),
        ("เลขที่เอกสาร", doc_no),
    ]
    r = 4
    for label, value in info_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        vc = ws.cell(row=r, column=3, value=value)
        vc.alignment = LEFT_WRAP
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=ncols)
        r += 1

    # ── ตารางรายการสินค้า ──
    header_row = r + 1
    for col, title in enumerate(STATEMENT_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BOX
    ws.row_dimensions[header_row].height = 20

    first_item_row = header_row + 1
    rr = first_item_row
    for idx, item in enumerate(cat_row.items, start=1):
        ws.cell(row=rr, column=1, value=idx).alignment = CENTER
        ws.cell(row=rr, column=2, value=item.sku)
        ws.cell(row=rr, column=3, value=item.item_name).alignment = LEFT_WRAP
        ws.cell(row=rr, column=4, value=item.variant_name)
        ws.cell(row=rr, column=5, value=item.qty_sold).number_format = QTY_FMT
        ws.cell(row=rr, column=6, value=item.unit_price).number_format = MONEY_FMT
        ws.cell(row=rr, column=7, value=item.gross_sales).number_format = MONEY_FMT
        ws.cell(row=rr, column=8, value=item.discounts).number_format = MONEY_FMT
        ws.cell(row=rr, column=9, value=item.net_sales).number_format = MONEY_FMT
        for c in range(1, ncols + 1):
            ws.cell(row=rr, column=c).border = BOX
        rr += 1
    last_item_row = rr - 1
    has_items = last_item_row >= first_item_row

    # แถวรวมท้ายตาราง — ใช้สูตรเพื่อให้ผู้รับตรวจยอดเองได้
    table_total_row = rr
    tc = ws.cell(row=table_total_row, column=1, value="รวม")
    tc.font = BOLD
    tc.alignment = CENTER
    ws.merge_cells(start_row=table_total_row, start_column=1, end_row=table_total_row, end_column=4)
    for col, fmt in ((5, QTY_FMT), (7, MONEY_FMT), (8, MONEY_FMT), (9, MONEY_FMT)):
        letter = get_column_letter(col)
        value = f"=SUM({letter}{first_item_row}:{letter}{last_item_row})" if has_items else 0
        cell = ws.cell(row=table_total_row, column=col, value=value)
        cell.font = BOLD
        cell.number_format = fmt
    for c in range(1, ncols + 1):
        ws.cell(row=table_total_row, column=c).border = BOX

    if has_items:
        ws.freeze_panes = ws.cell(row=first_item_row, column=1).coordinate

    # ── บล็อกสรุปยอด (ชิดขวาแบบ invoice) ──
    money_col = ncols            # I
    label_start_col = ncols - 3  # F
    s = table_total_row + 2

    def summary_line(row_i: int, label: str, value, *, fmt=MONEY_FMT,
                     label_font=None, value_font=None, border=None):
        lc = ws.cell(row=row_i, column=label_start_col, value=label)
        lc.font = label_font or Font(bold=True, size=10)
        lc.alignment = RIGHT
        ws.merge_cells(start_row=row_i, start_column=label_start_col,
                       end_row=row_i, end_column=money_col - 1)
        vc = ws.cell(row=row_i, column=money_col, value=value)
        vc.number_format = fmt
        vc.alignment = RIGHT
        if value_font:
            vc.font = value_font
        if border:
            lc.border = border
            vc.border = border
        return row_i + 1

    gross_ref = f"{get_column_letter(7)}{table_total_row}"
    disc_ref = f"{get_column_letter(8)}{table_total_row}"
    net_ref = f"{get_column_letter(9)}{table_total_row}"

    s = summary_line(s, "ยอดขายรวม (Gross)", f"={gross_ref}" if has_items else 0)
    s = summary_line(s, "หักส่วนลด", f"=-{disc_ref}" if has_items else 0)

    resid = _residual(cat_row.gross_sales, cat_row.discounts, cat_row.net_sales)
    if abs(resid) > 0.005:
        # ปกติไม่เกิดขึ้น (ร้านยังไม่มี VAT) — แต่ถ้าเกิด ต้องโชว์ ไม่ใช่ให้ยอดดูไม่ลงตัว
        s = summary_line(s, "ภาษี / รายการปรับปรุงอื่น", resid)

    s = summary_line(
        s, "ยอดสุทธิที่ต้องจ่าย", f"={net_ref}" if has_items else 0,
        label_font=PAYOUT_FONT, value_font=PAYOUT_FONT, border=DOUBLE_TOP,
    )
    s = summary_line(s, "จำนวนชิ้นที่ขายรวม",
                     f"=SUM({get_column_letter(5)}{first_item_row}:{get_column_letter(5)}{last_item_row})"
                     if has_items else 0, fmt=QTY_FMT)
    s = summary_line(s, "จำนวนบิลที่เกี่ยวข้อง", cat_row.receipts_count, fmt="#,##0")

    # ── หมายเหตุ + ช่องลงนาม ──
    note_row = s + 1
    notes = [
        f"หมายเหตุ: ยอดข้างต้นเป็นสกุลเงิน {config.REPORT_CURRENCY_SYMBOL} "
        f"หักรายการคืนสินค้า (refund) ออกแล้ว และไม่นับบิลที่ถูกยกเลิก",
        "ราคา/หน่วย เป็นค่าเฉลี่ยถ่วงน้ำหนัก หากสินค้ารายการเดียวกันถูกขายหลายราคาในช่วงเวลานี้",
    ]
    for text in notes:
        c = ws.cell(row=note_row, column=1, value=text)
        c.font = SUBTITLE_FONT
        c.alignment = LEFT_WRAP
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
        note_row += 1

    sign_row = note_row + 2
    ws.cell(row=sign_row, column=2, value="ผู้จ่ายเงิน ................................").font = SUBTITLE_FONT
    ws.merge_cells(start_row=sign_row, start_column=2, end_row=sign_row, end_column=4)
    ws.cell(row=sign_row, column=6, value="ผู้รับเงิน ................................").font = SUBTITLE_FONT
    ws.merge_cells(start_row=sign_row, start_column=6, end_row=sign_row, end_column=ncols)

    date_row = sign_row + 2
    ws.cell(row=date_row, column=2, value="วันที่ ................................").font = SUBTITLE_FONT
    ws.merge_cells(start_row=date_row, start_column=2, end_row=date_row, end_column=4)
    ws.cell(row=date_row, column=6, value="วันที่ ................................").font = SUBTITLE_FONT
    ws.merge_cells(start_row=date_row, start_column=6, end_row=date_row, end_column=ncols)

    # ── ตั้งค่าการพิมพ์: พอดี 1 หน้ากว้าง + ซ้ำหัวตารางทุกหน้า ──
    ws.print_area = f"A1:{last_col_letter}{date_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"


def build_workbook(report: SalesReport) -> io.BytesIO:
    """
    สร้าง workbook: ชีตสรุปรวม 1 ชีต + ชีต statement 1 ชีตต่อผู้ฝากขาย 1 ราย
    คืนเป็น BytesIO (พร้อมส่งผ่าน Flask send_file หรือเขียนไฟล์)
    """
    wb = Workbook()
    _write_overview_sheet(wb, report)

    used_names = {"สรุปรวมทุกราย"}
    for i, cat_row in enumerate(report.rows, start=1):
        sheet_name = _safe_sheet_name(cat_row.category_name, used_names)
        _write_statement_sheet(wb, report, cat_row, sheet_name, i)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
