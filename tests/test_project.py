"""
test_project.py -- Test suite for the loyverse package
Run from project root:  python tests/test_project.py
Tests run without real Google Sheets or Loyverse API credentials.
"""
import sys
import os
import tempfile
import inspect
import py_compile
from datetime import date
from pathlib import Path
from unittest.mock import patch, MagicMock

# Make the project root importable regardless of where this is run from.
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

PASS = "[PASS]"
FAIL = "[FAIL]"
results = []


def test(name, fn):
    try:
        fn()
        results.append((PASS, name))
        print(f"  {PASS}  {name}")
    except Exception as e:
        results.append((FAIL, name))
        print(f"  {FAIL}  {name}")
        print(f"       -> {e}")


def _reset(*modules):
    """Remove cached modules to force fresh reload"""
    for m in modules:
        sys.modules.pop(m, None)


# Base env — only truly required keys (STORE_ID is now optional)
BASE_ENV = {
    "LOYVERSE_TOKEN":          "TEST_TOKEN",
    "INPUT_SHEET_URL":         "https://docs.google.com/spreadsheets/d/FAKE_IN/edit",
    "MAPPING_SHEET_URL":       "https://docs.google.com/spreadsheets/d/FAKE_MAP/edit",
    "TRANSACTION_OUTPUT_MODE": "csv",
    "TRANSACTION_CSV_PATH":    "output/transactions_test.csv",
}


def full_env(**extra) -> dict:
    """Merge BASE_ENV with any extra overrides for inner patches"""
    return {**BASE_ENV, **extra}


# ─── 0. Syntax Check ──────────────────────────────────────────
print("\n=== [0] Syntax Check ===")
py_files = [
    "loyverse/config.py",
    "loyverse/loyverse_api.py",
    "loyverse/barcode_gen.py",
    "loyverse/sku_generator.py",
    "loyverse/shared_logic.py",
    "loyverse/sheets/auth.py",
    "loyverse/sheets/reader.py",
    "loyverse/sheets/writer.py",
    "loyverse/steps/step1_barcode_gen.py",
    "loyverse/steps/step2_stock_update.py",
    "loyverse/steps/legacy_sync.py",
    "loyverse/web/app.py",
    "loyverse/reports/catalog_cache.py",
    "loyverse/reports/sales_by_category.py",
    "loyverse/reports/excel_export.py",
]
for f in py_files:
    def _check(f=f):
        py_compile.compile(str(ROOT / f), doraise=True)
    test(f"Syntax: {f}", _check)


# ─── 1. Config ────────────────────────────────────────────────
print("\n=== [1] Config ===")

def _load_config():
    _reset("loyverse.config")
    with patch.dict(os.environ, BASE_ENV):
        from loyverse import config
        assert config.LOYVERSE_TOKEN == "TEST_TOKEN"
        assert config.LOYVERSE_STORE_ID == ""   # optional — default empty
        assert config.SKU_DIGIT_PAD == 3
        assert config.OUTPUT_DIR == "output"
        assert config.TRANSACTION_OUTPUT_MODE == "csv"
        assert config.TRANSACTION_SHEET_NAME == "Transactions"
        assert config.MAPPING_SHEET_NAME == "Mapping"
test("config: loads all values, STORE_ID optional (defaults '')", _load_config)

def _store_id_optional():
    """Running without LOYVERSE_STORE_ID must NOT raise"""
    _reset("loyverse.config")
    env_no_store = {k: v for k, v in BASE_ENV.items() if k != "LOYVERSE_STORE_ID"}
    with patch.dict(os.environ, env_no_store, clear=True):
        from loyverse import config
        assert config.LOYVERSE_STORE_ID == ""
test("config: LOYVERSE_STORE_ID is optional (no EnvironmentError)", _store_id_optional)

def _missing_token():
    """_require() must raise EnvironmentError when key is absent."""
    _reset("loyverse.config")
    original_getenv = os.getenv
    def fake_getenv(key, default=None):
        if key == "LOYVERSE_TOKEN":
            return None   # simulate missing
        return original_getenv(key, default)
    with patch("os.getenv", side_effect=fake_getenv):
        token_backup = os.environ.pop("LOYVERSE_TOKEN", None)
        try:
            from loyverse import config  # noqa: F401
            raise AssertionError("Should have raised EnvironmentError")
        except EnvironmentError as e:
            assert "LOYVERSE_TOKEN" in str(e)
        finally:
            _reset("loyverse.config")
            if token_backup:
                os.environ["LOYVERSE_TOKEN"] = token_backup
test("config: raises EnvironmentError when LOYVERSE_TOKEN missing", _missing_token)


# ─── 2. SKU Generator ─────────────────────────────────────────
print("\n=== [2] SKU Generator (pure logic) ===")

_reset("loyverse.config", "loyverse.sku_generator", "loyverse.sheets.auth")
sys.modules["loyverse.sheets.auth"] = MagicMock()
with patch.dict(os.environ, BASE_ENV):
    from loyverse import config
    from loyverse import sku_generator

def _first_item():
    assert sku_generator.get_next_sku("MM", None) == "MM001"
test("get_next_sku: first item (None) -> MM001", _first_item)

def _increment():
    assert sku_generator.get_next_sku("MM", "MM005") == "MM006"
test("get_next_sku: MM005 -> MM006", _increment)

def _fd_prefix():
    assert sku_generator.get_next_sku("FD", "FD099") == "FD100"
test("get_next_sku: FD099 -> FD100", _fd_prefix)

def _three_digit_pad():
    result = sku_generator.get_next_sku("AB", None)
    assert result == "AB001", f"Expected AB001, got '{result}'"
test("get_next_sku: 3-digit zero-pad enforced", _three_digit_pad)

def _extract_num():
    assert sku_generator._extract_number("MM042", "MM") == 42
test("_extract_number: MM042 -> 42", _extract_num)

def _extract_none():
    assert sku_generator._extract_number("FD001", "MM") is None
test("_extract_number: FD001 with prefix MM -> None (no match)", _extract_none)

def _extract_case():
    assert sku_generator._extract_number("mm010", "MM") == 10
test("_extract_number: case-insensitive (mm010 matches MM) -> 10", _extract_case)


# ─── 3. Sheets Auth ───────────────────────────────────────────
print("\n=== [3] Sheets Auth ===")

_reset("loyverse.sheets.auth", "loyverse.config")
with patch.dict(os.environ, BASE_ENV):
    from loyverse import config as _cfg  # noqa: F401

def _no_creds():
    _reset("loyverse.sheets.auth")
    from loyverse.sheets import auth as sheets_auth
    sheets_auth.get_client.cache_clear()
    try:
        sheets_auth.get_client()
        raise AssertionError("Should have raised FileNotFoundError")
    except FileNotFoundError as e:
        assert "credentials" in str(e).lower() or "ไม่พบ" in str(e)
test("sheets.auth: FileNotFoundError when credentials.json absent", _no_creds)


# ─── 4. Barcode Generator ─────────────────────────────────────
print("\n=== [4] Barcode Generator ===")

_reset("loyverse.barcode_gen")
from loyverse import barcode_gen

def _empty_sku():
    result = barcode_gen.generate("", Path(tempfile.mkdtemp()))
    assert result is None
test("barcode_gen: empty SKU -> None", _empty_sku)

def _safe_fn():
    result = barcode_gen._safe_filename('AB:CD/EF*GH?"<>|')
    for bad_char in [':', '/', '*', '?', '"', '<', '>', '|']:
        assert bad_char not in result, f"Illegal char '{bad_char}' still in: {result}"
test("barcode_gen: _safe_filename removes all illegal characters", _safe_fn)

def _gen_png():
    with tempfile.TemporaryDirectory() as tmp:
        path = barcode_gen.generate("TEST001", Path(tmp))
        assert path is not None
        assert path.endswith(".png")
        assert os.path.exists(path)
test("barcode_gen: creates PNG file for valid SKU (TEST001)", _gen_png)


# ─── 5. Transaction Writer (CSV mode) ─────────────────────────
print("\n=== [5] Transaction Writer (CSV mode) ===")

SAMPLE_RECORDS = [
    {
        "Product Name": "Test Product A", "SKU": "MM001", "Category": "Shoes",
        "Quantity Added": 5, "Stock Before": 10, "Stock After": 15,
        "Action": "Update Stock (Found by SKU)", "Status": "Success", "Detail/Error": "",
    },
    {
        "Product Name": "Test Product B", "SKU": "MM002", "Category": "Shoes",
        "Quantity Added": 3, "Stock Before": 0, "Stock After": 3,
        "Action": "Created New Item (Auto SKU: MM002)", "Status": "Success", "Detail/Error": "",
    },
    {
        "Product Name": "Bad Product", "SKU": "-", "Category": "Unknown",
        "Quantity Added": 1, "Stock Before": "-", "Stock After": "-",
        "Action": "", "Status": "Error", "Detail/Error": "Category not found",
    },
]


def _write_with_csv(records, csv_path, timestamp):
    env = full_env(TRANSACTION_CSV_PATH=str(csv_path), TRANSACTION_OUTPUT_MODE="csv")
    with patch.dict(os.environ, env):
        _reset("loyverse.config", "loyverse.sheets.writer")
        from loyverse import config as _c  # noqa: F401
        from loyverse.sheets import writer as sw
        sw.write_results(records, timestamp)


def _csv_creates_file():
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = Path(tmp) / "tx.csv"
        _write_with_csv(SAMPLE_RECORDS, csv_path, "20260311_130000")
        assert csv_path.exists(), "CSV file was not created"
test("sheets.writer (csv): creates CSV file", _csv_creates_file)

def _csv_has_header():
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = Path(tmp) / "tx.csv"
        _write_with_csv(SAMPLE_RECORDS, csv_path, "20260311_130000")
        with open(csv_path, encoding="utf-8-sig") as f:
            first_line = f.readline().strip()
        assert "Timestamp" in first_line and "Product Name" in first_line, \
            f"Header missing: {first_line}"
test("sheets.writer (csv): header row written correctly", _csv_has_header)

def _csv_row_count():
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = Path(tmp) / "tx.csv"
        _write_with_csv(SAMPLE_RECORDS, csv_path, "20260311_130000")
        with open(csv_path, encoding="utf-8-sig") as f:
            lines = f.readlines()
        assert len(lines) == 4, f"Expected 4 lines (header+3 data), got {len(lines)}"
test("sheets.writer (csv): correct row count (header + 3 data)", _csv_row_count)

def _csv_append():
    """Second write_results call appends, does not overwrite"""
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = Path(tmp) / "tx.csv"
        _write_with_csv(SAMPLE_RECORDS[:1], csv_path, "20260311_130000")
        _write_with_csv(SAMPLE_RECORDS[1:], csv_path, "20260311_140000")
        with open(csv_path, encoding="utf-8-sig") as f:
            lines = f.readlines()
        # header(1) + run1(1) + run2(2) = 4 lines
        assert len(lines) == 4, f"Expected 4 lines on append, got {len(lines)}"
test("sheets.writer (csv): appends on second run (no overwrite)", _csv_append)

def _csv_error_row_included():
    """Error rows must also be written to CSV"""
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = Path(tmp) / "tx.csv"
        _write_with_csv(SAMPLE_RECORDS, csv_path, "20260311_130000")
        with open(csv_path, encoding="utf-8-sig") as f:
            content = f.read()
        assert "Error" in content, "Error status not found in CSV"
        assert "Category not found" in content
test("sheets.writer (csv): error rows included in CSV", _csv_error_row_included)


# ─── 6. loyverse_api ──────────────────────────────────────────
print("\n=== [6] loyverse_api: structure & STORE_ID handling ===")

_reset("loyverse.loyverse_api", "loyverse.config")
sys.modules["loyverse.sheets.auth"] = MagicMock()
with patch.dict(os.environ, BASE_ENV):
    from loyverse import config as _cfg2  # noqa: F401
    from loyverse import loyverse_api as lv

def _fn_exists():
    required = [
        "find_variant_by_sku", "find_variant_by_name",
        "get_all_categories", "find_category_id",
        "get_last_sku_in_category", "get_current_stock",
        "update_stock", "create_item",
        "iter_all_items", "iter_receipts", "list_categories",
    ]
    for fn in required:
        assert hasattr(lv, fn), f"Missing function: {fn}"
test("loyverse_api: all required functions exist", _fn_exists)


def _iter_receipts_pagination():
    """iter_receipts() ต้องไล่ทุกหน้าตาม cursor และหยุดเมื่อไม่มี cursor (ไม่วน infinite)"""
    pages = [
        {"receipts": [{"receipt_number": "1"}, {"receipt_number": "2"}], "cursor": "PAGE2"},
        {"receipts": [{"receipt_number": "3"}]},
    ]
    calls = []

    def mock_get(path, params=None):
        calls.append(params)
        return pages.pop(0)

    with patch.object(lv, "_get", side_effect=mock_get), \
         patch.object(lv, "_get_default_store_id", return_value="STORE1"):
        result = list(lv.iter_receipts("2026-07-01T00:00:00.000Z", "2026-07-01T23:59:59.999Z"))

    assert [r["receipt_number"] for r in result] == ["1", "2", "3"]
    assert len(calls) == 2
    assert "cursor" not in calls[0]
    assert calls[1]["cursor"] == "PAGE2"
test("loyverse_api: iter_receipts paginates via cursor and stops correctly", _iter_receipts_pagination)

def _create_item_sig():
    sig = inspect.signature(lv.create_item)
    params = list(sig.parameters.keys())
    assert "category_id" in params, f"create_item missing category_id: {params}"
test("loyverse_api: create_item accepts category_id parameter", _create_item_sig)

def _update_stock_no_store_id():
    """update_stock payload must NOT include store_id when STORE_ID is empty"""
    captured = {}
    def mock_post(path, payload):
        captured["payload"] = payload
        mock_resp = MagicMock()
        mock_resp.raise_for_status = lambda: None
        return mock_resp

    _reset("loyverse.loyverse_api", "loyverse.config")
    with patch.dict(os.environ, full_env(LOYVERSE_STORE_ID="")):
        from loyverse import config as _c3  # noqa: F401
        from loyverse import loyverse_api as lv3
        with patch.object(lv3, "_post", side_effect=mock_post), \
             patch.object(lv3, "get_current_stock", return_value=5.0):
            lv3.update_stock("FAKE_VARIANT", 3)
    level = captured["payload"]["inventory_levels"][0]
    assert "store_id" not in level, f"store_id should be absent when empty, got: {level}"
test("loyverse_api: update_stock omits store_id when LOYVERSE_STORE_ID is empty", _update_stock_no_store_id)

def _create_item_no_store_id():
    """create_item store entry must NOT include store_id when STORE_ID is empty"""
    captured = {}
    def mock_post(path, payload):
        captured["payload"] = payload
        mock_resp = MagicMock()
        mock_resp.status_code = 201
        return mock_resp

    _reset("loyverse.loyverse_api", "loyverse.config")
    with patch.dict(os.environ, full_env(LOYVERSE_STORE_ID="")):
        from loyverse import config as _c4  # noqa: F401
        from loyverse import loyverse_api as lv4
        with patch.object(lv4, "_post", side_effect=mock_post):
            lv4.create_item("Test Item", "MM001", 5)
    store_entry = captured["payload"]["variants"][0]["stores"][0]
    assert "store_id" not in store_entry, f"store_id should be absent, got: {store_entry}"
    assert store_entry["stock_after"] == 5
test("loyverse_api: create_item omits store_id when LOYVERSE_STORE_ID is empty", _create_item_no_store_id)


# ─── 7. Reports: sales_by_category (pure aggregate logic) ─────
print("\n=== [7] Reports: sales_by_category ===")

_reset("loyverse.reports.catalog_cache", "loyverse.reports.sales_by_category", "loyverse.config")
with patch.dict(os.environ, BASE_ENV):
    from loyverse import config as _cfg3  # noqa: F401
    from loyverse.reports.catalog_cache import Catalog, ItemInfo, NO_CATEGORY_NAME
    from loyverse.reports import sales_by_category as sbc
    from loyverse.reports.sales_by_category import DELETED_ITEM_CATEGORY_NAME


def _make_catalog():
    items = {
        "item1": ItemInfo(item_id="item1", item_name="สินค้า A", category_id="catA", category_name="ร้านเอ"),
        "item2": ItemInfo(item_id="item2", item_name="สินค้า B", category_id="catB", category_name="ร้านบี"),
    }
    return Catalog(items=items, categories=[("catA", "ร้านเอ"), ("catB", "ร้านบี")])


def _sale_receipt(number, item_id, sku, qty, gross, discount, net, variant_id=None,
                  receipt_type="SALE", cancelled_at=None, price=None):
    return {
        "receipt_number": number,
        "receipt_type": receipt_type,
        "cancelled_at": cancelled_at,
        "line_items": [{
            "item_id": item_id,
            "variant_id": variant_id or f"{item_id}-v1",
            "sku": sku,
            "item_name": f"item-{item_id}",
            "variant_name": None,
            "quantity": qty,
            "price": price if price is not None else (gross / qty if qty else 0),
            "gross_total_money": gross,
            "total_discount": discount,
            "total_money": net,
        }],
    }


def _two_categories_share_100():
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18),
        _sale_receipt("1-002", "item2", "SKU2", 1, 15, 0, 15),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    assert len(report.rows) == 2
    by_name = {r.category_name: r for r in report.rows}
    assert by_name["ร้านเอ"].net_sales == 18
    assert by_name["ร้านบี"].net_sales == 15
    total_share = sum(r.share_pct for r in report.rows)
    assert abs(total_share - 100.0) < 1e-9, f"share_pct should sum to 100, got {total_share}"
    assert report.totals.net_sales == 33
test("aggregate: 2 SALE receipts / 2 categories -> correct totals, share_pct sums to 100", _two_categories_share_100)


def _refund_is_subtracted():
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18),
        _sale_receipt("1-002", "item1", "SKU1", 1, 10, 0, 9, receipt_type="REFUND"),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    row = report.rows[0]
    assert row.qty_sold == 1        # 2 - 1
    assert row.net_sales == 9       # 18 - 9
    assert report.refunds_scanned == 1
    assert report.receipts_scanned == 2
test("aggregate: REFUND receipt subtracts qty/net_sales, refunds_scanned counted", _refund_is_subtracted)


def _cancelled_receipt_skipped():
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18),
        _sale_receipt("1-002", "item1", "SKU1", 1, 10, 0, 10, cancelled_at="2026-07-01T10:00:00.000Z"),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    assert report.cancelled_skipped == 1
    assert report.receipts_scanned == 1
    assert report.rows[0].net_sales == 18   # ใบที่ยกเลิกไม่ถูกนับ
test("aggregate: cancelled receipt is skipped and counted in cancelled_skipped", _cancelled_receipt_skipped)


def _deleted_item_bucket():
    receipts = [_sale_receipt("1-001", "item_gone", "SKUX", 1, 10, 0, 10)]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    assert len(report.rows) == 1
    assert report.rows[0].category_name == DELETED_ITEM_CATEGORY_NAME
test("aggregate: item_id missing from catalog -> '(สินค้าถูกลบ)' bucket", _deleted_item_bucket)


def _category_filter_subset():
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18),
        _sale_receipt("1-002", "item2", "SKU2", 1, 15, 0, 15),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
        category_ids=["catA"],
    )
    assert len(report.rows) == 1
    assert report.rows[0].category_name == "ร้านเอ"
    assert report.rows[0].share_pct == 100.0
    assert report.totals.net_sales == 18
test("aggregate: category_ids filter narrows rows and recomputes share_pct on subset", _category_filter_subset)


def _fractional_qty_sold_by_weight():
    receipts = [_sale_receipt("1-001", "item1", "SKU1", 1.5, 15, 0, 15)]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    assert report.rows[0].qty_sold == 1.5
test("aggregate: sold_by_weight fractional quantity (1.5) preserved as float", _fractional_qty_sold_by_weight)


def _item_level_fields():
    """ItemRow ต้องมี gross/discount/ราคาต่อหน่วย ครบสำหรับทำ statement แบบ invoice"""
    receipts = [_sale_receipt("1-001", "item1", "SKU1", 4, 40, 4, 36)]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    item = report.rows[0].items[0]
    assert item.qty_sold == 4
    assert item.gross_sales == 40
    assert item.discounts == 4
    assert item.net_sales == 36
    assert item.unit_price == 10.0, item.unit_price   # 40 / 4
test("aggregate: ItemRow carries gross/discount/unit_price for invoice layout", _item_level_fields)


def _unit_price_weighted_average():
    """ขายสินค้าเดียวกันสองราคา → ราคา/หน่วย เป็นค่าเฉลี่ยถ่วงน้ำหนัก ไม่ใช่ราคาล่าสุด"""
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 1, 10, 0, 10, price=10),
        _sale_receipt("1-002", "item1", "SKU1", 3, 60, 0, 60, price=20),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    item = report.rows[0].items[0]
    assert item.qty_sold == 4
    assert item.unit_price == 17.5, item.unit_price   # (10 + 60) / 4
test("aggregate: unit_price is weighted average across differing prices", _unit_price_weighted_average)


def _unit_price_fallback_when_qty_zero():
    """ขายแล้วคืนหมด (qty สุทธิ = 0) ต้องไม่หารด้วยศูนย์ — ใช้ราคาตั้งเป็น fallback"""
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 0, 20, price=10),
        _sale_receipt("1-002", "item1", "SKU1", 2, 20, 0, 20, price=10, receipt_type="REFUND"),
    ]
    report = sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 1), tz_name="Asia/Bangkok",
    )
    item = report.rows[0].items[0]
    assert item.qty_sold == 0
    assert item.unit_price == 10.0, item.unit_price
test("aggregate: unit_price falls back to list price when net qty is 0 (no ZeroDivision)", _unit_price_fallback_when_qty_zero)


def _utc_range_bangkok():
    lo, hi = sbc.utc_range(date(2026, 7, 1), date(2026, 7, 1), "Asia/Bangkok")
    assert lo == "2026-06-30T17:00:00.000Z", lo
    assert hi == "2026-07-01T16:59:59.999Z", hi
test("utc_range: Asia/Bangkok single day -> correct UTC boundaries", _utc_range_bangkok)


def _catalog_cache_ttl_and_refresh():
    from loyverse.reports import catalog_cache as cc
    cc.clear_cache()
    fake_catalog = _make_catalog()
    with patch.object(cc, "_fetch_catalog", return_value=fake_catalog) as mock_fetch:
        c1 = cc.get_catalog()
        c2 = cc.get_catalog()  # ต้องมาจาก cache ไม่ยิงซ้ำ
        assert mock_fetch.call_count == 1
        c3 = cc.get_catalog(force_refresh=True)  # บังคับ refresh ต้องยิงใหม่
        assert mock_fetch.call_count == 2
    cc.clear_cache()
test("catalog_cache: caches within TTL, force_refresh bypasses cache", _catalog_cache_ttl_and_refresh)


# ─── 8. Reports: excel_export ──────────────────────────────────
print("\n=== [8] Reports: excel_export ===")

from loyverse.reports import excel_export


def _sample_report():
    receipts = [
        _sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18),
        _sale_receipt("1-002", "item2", "SKU2", 1, 15, 0, 15),
    ]
    return sbc.aggregate(
        receipts, _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 30), tz_name="Asia/Bangkok",
    )


def _excel_overview_plus_statement_per_owner():
    """1 ชีตสรุปรวม + 1 ชีต statement ต่อผู้ฝากขาย 1 ราย"""
    import openpyxl
    report = _sample_report()
    bio = excel_export.build_workbook(report)
    wb = openpyxl.load_workbook(bio)
    assert wb.sheetnames == ["สรุปรวมทุกราย", "ร้านเอ", "ร้านบี"], wb.sheetnames

    ws = wb["สรุปรวมทุกราย"]
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=12):
        if [c.value for c in row][:1] == [excel_export.SUMMARY_HEADERS[0]]:
            header_row = row[0].row
            break
    assert header_row, "ไม่พบแถว header ในชีตสรุปรวม"
    assert [c.value for c in ws[header_row]] == excel_export.SUMMARY_HEADERS
test("excel_export: overview sheet + one statement sheet per consignor", _excel_overview_plus_statement_per_owner)


def _excel_statement_line_items():
    """ชีต statement ต้องมีหัวตารางแบบ invoice และรายการสินค้าพร้อมราคา/หน่วย"""
    import openpyxl
    report = _sample_report()
    bio = excel_export.build_workbook(report)
    ws = openpyxl.load_workbook(bio)["ร้านเอ"]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        if [c.value for c in row][:1] == [excel_export.STATEMENT_HEADERS[0]]:
            header_row = row[0].row
            break
    assert header_row, "ไม่พบแถว header ในชีต statement"
    assert [c.value for c in ws[header_row]][:9] == excel_export.STATEMENT_HEADERS

    # แถวสินค้าแรก: SKU1, qty 2, ราคา/หน่วย 10 (gross 20 / qty 2), net 18
    item = ws[header_row + 1]
    assert item[1].value == "SKU1", item[1].value
    assert item[4].value == 2
    assert item[5].value == 10.0, item[5].value
    assert item[8].value == 18

    # ต้องมีบล็อกสรุปที่ระบุยอดต้องจ่าย
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "ยอดสุทธิที่ต้องจ่าย" in text
    assert "ผู้รับเงิน" in text
test("excel_export: statement sheet has invoice line items + payout block", _excel_statement_line_items)


def _excel_safe_sheet_name():
    used = set()
    assert excel_export._safe_sheet_name("ร้าน/เอ:ทดสอบ", used) == "ร้าน-เอ-ทดสอบ"
    long_name = "ก" * 40
    assert len(excel_export._safe_sheet_name(long_name, used)) <= 31
    used2 = set()
    a = excel_export._safe_sheet_name("ซ้ำ", used2)
    b = excel_export._safe_sheet_name("ซ้ำ", used2)
    assert a != b, "ชื่อชีตซ้ำต้องถูกทำให้ไม่ซ้ำ"
test("excel_export: _safe_sheet_name strips illegal chars, truncates, dedupes", _excel_safe_sheet_name)


def _excel_filename():
    report = _sample_report()
    fname = excel_export.suggested_filename(report)
    assert fname == "สรุปยอดฝากขาย_ทุกราย_20260701-20260730.xlsx", fname

    # ผู้ฝากขายรายเดียว → ใส่ชื่อรายนั้นในชื่อไฟล์ ส่งต่อได้เลย
    single = sbc.aggregate(
        [_sale_receipt("1-001", "item1", "SKU1", 2, 20, 2, 18)], _make_catalog(),
        date_from=date(2026, 7, 1), date_to=date(2026, 7, 30), tz_name="Asia/Bangkok",
    )
    assert excel_export.suggested_filename(single) == "สรุปยอดฝากขาย_ร้านเอ_20260701-20260730.xlsx"
test("excel_export: suggested_filename names the consignor when report has one", _excel_filename)


# ─── 9. Web: /reports/sales routes ─────────────────────────────
print("\n=== [9] Web: /reports/sales routes ===")

_reset("loyverse.web.app", "loyverse.config", "loyverse.sheets.auth")
sys.modules["loyverse.sheets.auth"] = MagicMock()
with patch.dict(os.environ, BASE_ENV):
    from loyverse import config as _cfg4  # noqa: F401
    from loyverse.web import app as web_app


def _fake_catalog_for_web():
    from loyverse.reports.catalog_cache import Catalog
    return Catalog(items={}, categories=[("catA", "ร้านเอ"), ("catB", "ร้านบี")])


def _fake_build_report(date_from, date_to, category_ids=None, progress=None):
    from datetime import datetime, timezone
    row = sbc.CategoryRow(category_id="catA", category_name="ร้านเอ", qty_sold=2, gross_sales=20,
                          discounts=2, net_sales=18, share_pct=100.0, receipts_count=1, items=[])
    totals = sbc.CategoryRow(category_id=None, category_name="รวมทั้งหมด", qty_sold=2, gross_sales=20,
                             discounts=2, net_sales=18, share_pct=100.0, receipts_count=1)
    return sbc.SalesReport(
        date_from=date_from, date_to=date_to, tz_name="Asia/Bangkok",
        category_filter=category_ids or [],
        rows=[row], totals=totals,
        receipts_scanned=1, refunds_scanned=0, cancelled_skipped=0,
        generated_at=datetime.now(timezone.utc),
    )


def _sales_report_default_page():
    with patch.object(web_app.catalog_cache, "get_catalog", return_value=_fake_catalog_for_web()), \
         patch.object(web_app.sbc, "build_report", side_effect=_fake_build_report):
        client = web_app.app.test_client()
        res = client.get("/reports/sales")
    assert res.status_code == 200
    assert "ร้านเอ".encode() in res.data
test("web: GET /reports/sales (no params) -> 200 with default range + category", _sales_report_default_page)


def _sales_report_bad_range():
    with patch.object(web_app.catalog_cache, "get_catalog", return_value=_fake_catalog_for_web()):
        client = web_app.app.test_client()
        res = client.get("/reports/sales?from=2026-07-30&to=2026-07-01")
    assert res.status_code == 200
    assert "ช่วงวันที่ไม่ถูกต้อง".encode() in res.data
test("web: GET /reports/sales with to < from -> 200 with Thai error message (not 500)", _sales_report_bad_range)


def _sales_report_export():
    with patch.object(web_app.sbc, "build_report", side_effect=_fake_build_report):
        client = web_app.app.test_client()
        res = client.get("/reports/sales/export?from=2026-07-01&to=2026-07-30")
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in res.headers.get("Content-Disposition", "")
test("web: GET /reports/sales/export -> 200 xlsx attachment", _sales_report_export)


def _parse_report_args_edge_cases():
    from werkzeug.datastructures import MultiDict

    _, _, _, err = web_app._parse_report_args(MultiDict({"from": "not-a-date"}))
    assert err, "invalid date should produce an error message"

    _, _, cats, _ = web_app._parse_report_args(MultiDict([("category", "a"), ("category", "b")]))
    assert cats == ["a", "b"]

    today_str = date.today().isoformat()
    _, _, _, err2 = web_app._parse_report_args(MultiDict({"from": "2000-01-01", "to": today_str}))
    assert "กว้างเกินไป" in err2
test("web: _parse_report_args handles bad date / repeated category / too-wide range", _parse_report_args_edge_cases)


# ─── Summary ──────────────────────────────────────────────────
print("\n" + "=" * 55)
passed = sum(1 for r in results if r[0] == PASS)
total  = len(results)
print(f"  RESULT: {passed}/{total} tests passed")
if passed == total:
    print("  All tests PASSED!")
else:
    failed_names = [r[1] for r in results if r[0] == FAIL]
    print(f"  {len(failed_names)} test(s) FAILED:")
    for n in failed_names:
        print(f"    - {n}")
print("=" * 55)

sys.exit(0 if passed == total else 1)
